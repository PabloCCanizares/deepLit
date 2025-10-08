import os
import re
import io
import base64
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_pymongo import PyMongo
import PyPDF2
from bson.objectid import ObjectId

# Configurar Matplotlib para usar un backend no interactivo
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from wordcloud import WordCloud

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Necesario para mensajes flash

# Configuración de MongoDB (ajusta la URI según tu entorno)
app.config["MONGO_URI"] = "mongodb://localhost:27017/mydatabase"
mongo = PyMongo(app)

############################
# Funciones de extracción existentes
############################

def extract_text_from_pdf(pdf_path):
    """Extrae el texto completo de un archivo PDF."""
    full_text = ""
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
    return full_text

def extract_title(text):
    """Extrae el título del documento (se asume que es la primera línea significativa)."""
    for line in text.splitlines():
        cleaned = line.strip()
        if cleaned and len(cleaned) > 5:
            return cleaned
    return "No se encontró el título."

def extract_abstract(text):
    """
    Busca y extrae el abstract.
    Se asume que comienza con "Abstract" y termina antes de 'Keywords', 'Palabras clave', 'Introduction' o 'Introducción'.
    """
    pattern = r'(Abstract\s*[:.-]*)(.*?)(?=\n\s*(Keywords|Palabras clave|Introduction|Introducción))'
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return match.group(2).strip() if match else "No se encontró el abstract."

def extract_keywords(text):
    """
    Extrae los keywords a partir de "Keywords" o "Palabras clave:".
    """
    pattern = r'(Keywords|Palabras clave)\s*[:.-]*\s*(.*?)(?=\n)'
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(2).strip() if match else "No se encontraron keywords."

def extract_bibliography(text):
    """
    Extrae la bibliografía asumiendo que su encabezado (References, Bibliography, etc.)
    aparece al inicio de una línea o del documento.
    """
    pattern = r'(^|\n)\s*(References|Bibliography|eferences)\s*[:.-]*\s*(.*)'
    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE | re.DOTALL)
    if match:
        biblio_text = match.group(3).strip()
        end_pattern = r'(?:\n\s*(?:Appendix|Anexo|Apendices|Supplementary|Suplementario)|received the)'
        end_match = re.search(end_pattern, biblio_text, re.IGNORECASE)
        if end_match:
            biblio_text = biblio_text[:end_match.start()].strip()
        return biblio_text
    return "No se encontró la bibliografía."

def extract_citations(text):
    """
    Extrae todas las citas del bloque de bibliografía.
    Se asume que cada cita comienza con [número] y se extiende hasta la siguiente o el final.
    """
    pattern = re.compile(r'\[(\d+)\](.*?)(?=\[\d+\]|$)', re.DOTALL)
    citations = []
    for match in pattern.finditer(text):
        number = match.group(1)
        content = match.group(2)
        cleaned = ' '.join(content.split())
        citation_entry = f"[{number}] {cleaned}"
        citations.append(citation_entry)
    return citations

############################
# Nuevas funciones de extracción para campos adicionales
############################

def extract_anio(text):
    """Extrae el año (Año) del texto, buscando patrones como 'Año: 2022'."""
    pattern = r"Año\s*[:\-]\s*(\d{4})"
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(1) if match else ""

def extract_categoria(text):
    """Extrae la categoría (Categoría) del texto."""
    pattern = r"Categor[ií]a\s*[:\-]\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).split("\n")[0].strip()
    return ""

def extract_tipo(text):
    """Extrae el tipo (Tipo) del texto."""
    pattern = r"Tipo\s*[:\-]\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).split("\n")[0].strip()
    return ""

def extract_acronimo(text):
    """Extrae el acrónimo (Acrónimo) del texto."""
    pattern = r"Acr[oó]nimo\s*[:\-]\s*(\w+)"
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def extract_paginas(text):
    """Extrae el número de páginas (Pág.) del texto."""
    pattern = r"Pág\.?\s*[:\-]?\s*(\d+)"
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def extract_obs(text):
    """Extrae las observaciones (Obs) del texto."""
    pattern = r"Obs\s*[:\-]\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).split("\n")[0].strip()
    return ""

def extract_resumen(text):
    """Extrae el resumen (Resumen) del texto."""
    pattern = r"Resumen\s*[:\-]\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return match.group(1).strip() if match else ""

def extract_enlace(text):
    """Extrae el enlace (Enlace) del texto."""
    pattern = r"Enlace\s*[:\-]\s*(\S+)"
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def extract_cita(text):
    """Extrae una cita bibliográfica (Cita) del texto."""
    pattern = r"Cita\s*[:\-]\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return match.group(1).strip() if match else ""

############################
# Función para generar wordcloud
############################

def generate_wordcloud():
    """
    Genera una wordcloud a partir de las keywords de todos los documentos en la base de datos.
    Devuelve la imagen codificada en base64 para incrustarla en la web.
    """
    docs = mongo.db.documents.find()
    text = ""
    for doc in docs:
        if "keywords" in doc:
            text += " " + doc["keywords"]
    if not text.strip():
        text = "No keywords"
    wc = WordCloud(width=800, height=200, background_color="white").generate(text)
    plt.figure(figsize=(8, 2))
    plt.imshow(wc, interpolation="bilinear")
    plt.axis("off")
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close()  # Liberar memoria
    buf.seek(0)
    img_base64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return img_base64

############################
# Rutas de la aplicación
############################

@app.route('/')
def dashboard():
    """Dashboard avanzado: estadísticas, diagrama de barras por año, ranking de keywords, actividad reciente, notificaciones y wordcloud."""
    docs = list(mongo.db.documents.find())
    total_documents = len(docs)
    total_references = sum(len(doc.get("citations", [])) for doc in docs)
    avg_references = round(total_references / total_documents, 2) if total_documents > 0 else 0

    # Diagrama de barras por año (usando el campo "Year")
    year_counts = {}
    for doc in docs:
        # Se asume que el campo "Year" viene como valor numérico o cadena convertible a int.
        if "Year" in doc and doc["Year"]:
            try:
                year = int(doc["Year"])
                year_counts[year] = year_counts.get(year, 0) + 1
            except ValueError:
                # Si no se puede convertir, se omite.
                pass
    sorted_years = sorted(year_counts.keys())
    chart_labels = [str(year) for year in sorted_years]
    chart_values = [year_counts[year] for year in sorted_years]

    # Ranking de keywords
    keyword_freq = {}
    for doc in docs:
        kw = doc.get("keywords", "")
        if kw and kw != "No se encontraron keywords":
            for word in kw.split(','):
                word = word.strip().lower()
                if word:
                    keyword_freq[word] = keyword_freq.get(word, 0) + 1
    sorted_keywords = sorted(keyword_freq.items(), key=lambda x: x[1], reverse=True)

    # Actividad reciente: los 5 documentos más recientes
    recent_docs = list(mongo.db.documents.find().sort("_id", -1).limit(5))
    for doc in recent_docs:
        oid = doc["_id"] if isinstance(doc["_id"], ObjectId) else ObjectId(doc["_id"])
        doc["upload_date"] = oid.generation_time.strftime("%Y-%m-%d %H:%M:%S")
        doc["_id"] = str(doc["_id"])

    # Notificaciones: documentos con errores de extracción
    notif_abstract = sum(1 for doc in docs if doc.get("abstract", "").strip() == "No se encontró el abstract.")
    notif_keywords = sum(1 for doc in docs if doc.get("keywords", "").strip() == "No se encontraron keywords.")

    # Generar la wordcloud
    wordcloud_img = generate_wordcloud()

    return render_template('dashboard.html',
                           total_documents=total_documents,
                           total_references=total_references,
                           avg_references=avg_references,
                           chart_labels=chart_labels,
                           chart_values=chart_values,
                           sorted_keywords=sorted_keywords,
                           recent_docs=recent_docs,
                           notif_abstract=notif_abstract,
                           notif_keywords=notif_keywords,
                           wordcloud_img=wordcloud_img)


@app.route('/upload_pdf_page')
def upload_pdf_page():
    """Página para subir un PDF."""
    return render_template('upload.html')

@app.route('/upload_pdf', methods=['POST'])
def upload_pdf():
    """Procesa la subida de un PDF, extrae la información (incluyendo campos adicionales) y la guarda en MongoDB."""
    if 'pdf' not in request.files:
        flash("No se encontró el archivo PDF.")
        return redirect(url_for('upload_pdf_page'))
    
    file = request.files['pdf']
    if file.filename == '':
        flash("No se seleccionó ningún archivo.")
        return redirect(url_for('upload_pdf_page'))
    
    # Opcional: los nuevos campos pueden enviarse manualmente desde el formulario,
    # o se intentará extraerlos automáticamente del texto del PDF.
    autores = request.form.get('autores', '').strip()  # Si se quiere agregar autores manualmente
    # Aquí podrías agregar más campos manuales si lo deseas.

    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)
    temp_path = os.path.join(temp_folder, file.filename)
    file.save(temp_path)
    
    text = extract_text_from_pdf(temp_path)
    title = extract_title(text)
    abstract = extract_abstract(text)
    keywords = extract_keywords(text)
    bibliography = extract_bibliography(text)
    citations = extract_citations(bibliography)
    
    # Nuevos campos extraídos automáticamente
    anio = extract_anio(text)
    categoria = extract_categoria(text)
    tipo = extract_tipo(text)
    acronimo = extract_acronimo(text)
    paginas = extract_paginas(text)
    obs = extract_obs(text)
    resumen = extract_resumen(text)
    enlace = extract_enlace(text)
    cita = extract_cita(text)
    
    doc = {
        "filename": file.filename,
        "title": title,
        "abstract": abstract,
        "keywords": keywords,
        "bibliography": bibliography,
        "citations": citations,
        # Campos adicionales
        "autores": autores,
        "anio": anio,
        "categoria": categoria,
        "tipo": tipo,
        "acronimo": acronimo,
        "paginas": paginas,
        "obs": obs,
        "resumen": resumen,
        "enlace": enlace,
        "cita": cita
    }
    mongo.db.documents.insert_one(doc)
    os.remove(temp_path)
    
    return render_template('result.html', document=doc)

@app.route('/home')
def home():
    # Carga la configuración; si no existe, usa valores por defecto
    config = mongo.db.configuration.find_one({"_id": "settings"})
    if not config:
        config = {
            "_id": "settings",
            "required_fields": ["Year", "Title", "Category", "Type", "Acronym", "Cites", "Pag.", "Obs.", "Summary", "link", "citation", "abstract"],
            "order_by_year": "asc"  # Valor por defecto
        }
        mongo.db.configuration.insert_one(config)
    else:
        # Si no existe "order_by_year" en config, se le asigna por defecto
        if "order_by_year" not in config:
            config["order_by_year"] = "asc"
            mongo.db.configuration.update_one({"_id": "settings"}, {"$set": {"order_by_year": "asc"}}, upsert=True)
    
    # Consulta todos los documentos
    documents_cursor = mongo.db.documents.find()
    documents = []
    for doc in documents_cursor:
        doc["_id"] = str(doc["_id"])
        documents.append(doc)

    def is_valid_year(val):
        try:
            int(val)
            return True
        except (ValueError, TypeError):
            return False

    # Separa documentos con año válido y los que no lo tienen
    valid_docs = [d for d in documents if is_valid_year(d.get("Year"))]
    invalid_docs = [d for d in documents if not is_valid_year(d.get("Year"))]

    # Ordena los documentos válidos según la configuración ("asc" o "desc")
    if config.get("order_by_year") == "desc":
        valid_docs = sorted(valid_docs, key=lambda d: int(d.get("Year")), reverse=True)
    else:
        valid_docs = sorted(valid_docs, key=lambda d: int(d.get("Year")))
    
    # Los documentos sin año válido se colocan al final
    sorted_docs = valid_docs + invalid_docs

    return render_template('home.html', documents=sorted_docs, config=config)



from scholarly import scholarly
import requests

def search_google_scholar(query):
    """Busca en Google Scholar usando scholarly"""
    search_results = scholarly.search_pubs(query)
    try:
        publication = next(search_results)
        return {
            "title": publication['bib']['title'],
            "authors": publication['bib']['author'],
            "year": publication['bib'].get('year', 'Desconocido'),
            "venue": publication['bib'].get('venue', 'Desconocido'),
            "num_citations": publication.get('num_citations', 'Desconocido'),
            "abstract": publication['bib'].get('abstract', 'No disponible'),
            "pub_url": publication.get('pub_url', 'No disponible'),
        }
    except StopIteration:
        return None

def search_semantic_scholar(query):
    """Busca en Semantic Scholar usando su API"""
    url = f"https://api.semanticscholar.org/graph/v1/paper/search?query={query}&fields=title,authors,year,venue,citationCount,abstract,url"
    response = requests.get(url)
    
    if response.status_code == 200:
        results = response.json().get('data', [])
        if results:
            paper = results[0]  # Tomamos el primer resultado
            return {
                "title": paper.get('title', 'Desconocido'),
                "authors": ", ".join([author['name'] for author in paper.get('authors', [])]),
                "year": paper.get('year', 'Desconocido'),
                "venue": paper.get('venue', 'Desconocido'),
                "num_citations": paper.get('citationCount', 'Desconocido'),
                "abstract": paper.get('abstract', 'No disponible'),
                "pub_url": paper.get('url', 'No disponible'),
            }
    return None

@app.route("/scholar", methods=["GET"])
def scholar():
    """Procesa la búsqueda y envía los datos a scholar_results.html"""
    query = "Testing the untestable"#request.form["article_title"]
    
    data = {
        "scholarly": search_google_scholar(query),
        "semantic": search_semantic_scholar(query)
    }
    
    return render_template("scholar_results.html", data=data)

@app.route('/result/<id>')
def result(id):
    """Muestra el detalle completo de un documento, incluyendo los nuevos campos."""
    doc = mongo.db.documents.find_one({"_id": ObjectId(id)})
    if doc:
        oid = doc["_id"] if isinstance(doc["_id"], ObjectId) else ObjectId(doc["_id"])
        doc["upload_date"] = oid.generation_time.strftime("%Y-%m-%d %H:%M:%S")
        doc["_id"] = str(doc["_id"])
        return render_template('result.html', document=doc)
    flash("Documento no encontrado")
    return redirect(url_for('home'))

@app.route('/search', methods=['GET', 'POST'])
def search():
    """Realiza la búsqueda de documentos por título."""
    if request.method == 'POST':
        query = request.form.get('query', '')
        if query:
            results_cursor = mongo.db.documents.find({"title": {"$regex": query, "$options": "i"}})
            results = []
            for doc in results_cursor:
                doc["_id"] = str(doc["_id"])
                results.append(doc)
        else:
            results = []
        return render_template('home.html', documents=results, query=query)
    return render_template('home.html', documents=None)

@app.route('/upload_folder', methods=['GET', 'POST'])
def upload_folder():
    """Sube una carpeta (o varios archivos PDF) y procesa cada documento."""
    if request.method == 'POST':
        files = request.files.getlist('pdfs')
        if not files or len(files) == 0:
            flash("No se seleccionaron archivos.")
            return redirect(url_for('upload_folder'))
        
        processed_docs = []
        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)
        
        for file in files:
            if file.filename == '' or not file.filename.lower().endswith('.pdf'):
                continue  # Omitir archivos vacíos o que no sean PDF
            temp_path = os.path.join(temp_folder, file.filename)
            file.save(temp_path)
            
            # Procesar el PDF usando las funciones existentes
            text = extract_text_from_pdf(temp_path)
            title = extract_title(text)
            abstract = extract_abstract(text)
            keywords = extract_keywords(text)
            bibliography = extract_bibliography(text)
            citations = extract_citations(bibliography)
            
            # También se intentan extraer los nuevos campos
            anio = extract_anio(text)
            categoria = extract_categoria(text)
            tipo = extract_tipo(text)
            acronimo = extract_acronimo(text)
            paginas = extract_paginas(text)
            obs = extract_obs(text)
            resumen = extract_resumen(text)
            enlace = extract_enlace(text)
            cita = extract_cita(text)
            
            doc = {
                "filename": file.filename,
                "title": title,
                "abstract": abstract,
                "keywords": keywords,
                "bibliography": bibliography,
                "citations": citations,
                # Campos adicionales
                "anio": anio,
                "categoria": categoria,
                "tipo": tipo,
                "acronimo": acronimo,
                "paginas": paginas,
                "obs": obs,
                "resumen": resumen,
                "enlace": enlace,
                "cita": cita
            }
            mongo.db.documents.insert_one(doc)
            processed_docs.append(doc)
            os.remove(temp_path)
        
        flash(f"Se han subido {len(processed_docs)} documentos.")
        return render_template('result_folder.html', documents=processed_docs)
    
    # GET: mostrar el formulario para subir la carpeta
    return render_template('upload_folder.html')

@app.route('/delete_all', methods=['POST'])
def delete_all():
    mongo.db.documents.delete_many({})
    flash("Se han borrado todas las entradas.")
    return redirect(url_for('dashboard'))

@app.route('/upload_excel', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        # Verificamos que el input del formulario sea 'excel'
        if 'excel' not in request.files:
            flash("No se encontró el archivo Excel.")
            return redirect(url_for('upload_excel'))
        
        file = request.files['excel']
        if file.filename == '':
            flash("No se seleccionó ningún archivo Excel.")
            return redirect(url_for('upload_excel'))
        
        # Guardar el archivo temporalmente
        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)
        temp_path = os.path.join(temp_folder, file.filename)
        file.save(temp_path)
        
        # Procesar el Excel
        processed_docs = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_path)
            sheet = wb.active  # Toma la primera hoja
            
            # Encabezados esperados (en el orden que quieras procesarlos)
            required_columns = [
                "Year", "Title", "Category", "Type", "Acronym", "Cites",
                "Pag.", "Obs.", "Summary", "link", "citation", "abstract"
            ]
            
            # Leer la primera fila (encabezados del Excel)
            headers = [cell.value for cell in sheet[1]]
            
            for cell in sheet[1]:
                print(cell.value) 
            # Comprobar que las columnas requeridas estén presentes
            for col_name in required_columns:
                if col_name not in headers:
                    flash(f"No se encontró la columna requerida: {col_name}.")
                    return redirect(url_for('upload_excel'))
            
            # Crear un mapa de nombre_columna -> índice (para acceder a cada valor de la fila)
            col_index_map = {}
            for col_name in required_columns:
                col_index_map[col_name] = headers.index(col_name) + 1  # +1 porque openpyxl es base 1
            
            # Recorrer las filas a partir de la 2 (la primera es encabezado)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Si la fila está completamente vacía, saltarla
                if not any(row):
                    continue
                
                # Construir el documento (diccionario)
                doc = {}
                for col_name in required_columns:
                    idx = col_index_map[col_name] - 1  # -1 porque row es base 0
                    valor = row[idx] if idx < len(row) else None
                    # Ajusta los nombres de campos en MongoDB según prefieras
                    # Aquí usamos directamente el nombre de la columna como clave
                    doc[col_name] = valor if valor is not None else ""
                
                # Insertar en MongoDB
                mongo.db.documents.insert_one(doc)
                processed_docs.append(doc)
            
            flash(f"Se han subido {len(processed_docs)} documentos desde el Excel.")
        except Exception as e:
            flash(f"Ocurrió un error procesando el Excel: {str(e)}")
        finally:
            # Eliminar el archivo temporal
            os.remove(temp_path)
        
        # Renderizamos una plantilla con el resultado
        return render_template('result_excel.html', documents=processed_docs)
    
    # GET: Mostrar el formulario para subir Excel
    return render_template('upload_excel.html')

@app.route('/config', methods=['GET', 'POST'])
def config():
    if request.method == 'POST':
        # Recoge el orden y los campos obligatorios seleccionados
        new_config = {
            "order_by_year": request.form.get("order_by_year", "asc"),
            "required_fields": request.form.getlist("required_fields")
        }
        # Actualiza (o inserta si no existe) la configuración en la colección "configuration"
        mongo.db.configuration.update_one({}, {"$set": new_config}, upsert=True)
        flash("Configuración actualizada correctamente")
        return redirect(url_for('config'))
    
    # En GET, recupera la configuración actual o usa valores por defecto
    config_data = mongo.db.configuration.find_one() or {"order_by_year": "asc", "required_fields": []}
    
    # Extraer los campos del esquema: usamos un documento de muestra de la colección documents
    sample_doc = mongo.db.documents.find_one() or {}
    # Excluimos campos que no nos interesan (por ejemplo, _id, upload_date, etc.)
    excluded = {"_id", "upload_date"}
    available_fields = [ key for key in sample_doc.keys() if key not in excluded ]
    
    # Opcionalmente, podemos ordenar alfabéticamente
    available_fields.sort()
    
    return render_template("config.html", config=config_data, available_fields=available_fields)

@app.route('/save_config', methods=['POST'])
def save_config():
    """
    Guarda la configuración de campos requeridos (u otros ajustes)
    en la colección 'configuration'. Usamos un documento único con _id="settings".
    """
    # Se espera que se envíe JSON con la clave "required_fields"
    data = request.get_json()
    required_fields = data.get("required_fields", [])
    
    config_doc = {"_id": "settings", "required_fields": required_fields}
    mongo.db.configuration.replace_one({"_id": "settings"}, config_doc, upsert=True)
    return {"status": "success"}, 200

@app.route('/save_sort_order', methods=['POST'])
def save_sort_order():
    data = request.get_json()
    sort_order = data.get("sort_order", "asc")
    # Guarda la configuración en la colección 'config'. Usamos un documento con _id "document_config"
    mongo.db.config.update_one(
        {"_id": "document_config"},
        {"$set": {"sort_order": sort_order}},
        upsert=True
    )
    return {"status": "success", "sort_order": sort_order}

if __name__ == '__main__':
    app.run(debug=True, port=5001)
