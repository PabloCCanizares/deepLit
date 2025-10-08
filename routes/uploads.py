############################################################
############################################################
    # RAMA MAIN #
############################################################
############################################################


# routes/uploads.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from bson.objectid import ObjectId
from extensions import mongo
from utils import (
    extract_text_from_pdf, extract_title, extract_abstract,
    extract_keywords, extract_bibliography, extract_citations,
    extract_anio, extract_categoria, extract_tipo, extract_acronimo,
    extract_paginas, extract_obs, extract_resumen, extract_enlace, extract_cita
)
import os
import zipfile
import tempfile
import uuid
import json
from routes.documents import edit_document as documents_edit_document

uploads_bp = Blueprint('upload', __name__, template_folder='../templates')


@uploads_bp.route('/upload', methods=['GET'])
def upload_page():
    """Página para subir un PDF."""
    return render_template('upload.html')

    
@uploads_bp.route('/upload_pdf', methods=['POST'])
def upload_pdf():
    if 'pdf' not in request.files:
        print("No se encontró el archivo PDF.")
        flash("No se encontró el archivo PDF.")
        return redirect(url_for('upload.upload_page'))
    
    file = request.files['pdf']
    if file.filename == '':
        print("No se seleccionó ningún archivo.")
        flash("No se seleccionó ningún archivo.")
        return redirect(url_for('upload.upload_page'))
    
    # autores = request.form.get('autores', '').strip()
    # anio = request.form.get('anio', '').strip()
    # categoria = request.form.get('categoria', '').strip()
    # tipo = request.form.get('tipo', '').strip()
    # acronimo = request.form.get('acronimo', '').strip()
    # paginas = request.form.get('paginas', '').strip()
    # obs = request.form.get('obs', '').strip()  
    # resumen = request.form.get('resumen', '').strip()
    # enlace = request.form.get('enlace', '').strip()
    # cita = request.form.get('cita', '').strip()

    # temp_folder = os.path.join("temp")
    # os.makedirs(temp_folder, exist_ok=True)
    # temp_path = os.path.join(temp_folder, file.filename)
    # file.save(temp_path)
    
    # text = extract_text_from_pdf(temp_path)
    # title = extract_title(text)
    # abstract = extract_abstract(text)
    # keywords = extract_keywords(text)
    # bibliography = extract_bibliography(text)
    # citations = extract_citations(bibliography)
    
    # anio = extract_anio(text) if not anio else anio
    # categoria = extract_categoria(text) if not categoria else categoria
    # tipo = extract_tipo(text) if not tipo else tipo
    # acronimo = extract_acronimo(text) if not acronimo else acronimo
    # paginas = extract_paginas(text) if not paginas else paginas
    # obs = extract_obs(text) if not obs else obs
    # resumen = extract_resumen(text) if not resumen else resumen
    # enlace = extract_enlace(text) if not enlace else enlace
    # cita = extract_cita(text) if not cita else cita

    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)
    temp_path = os.path.join(temp_folder, file.filename)
    file.save(temp_path)
    text = extract_text_from_pdf(temp_path)
    title = extract_title(text)
    abstract = extract_abstract(text)
    keywords = extract_keywords(text)
    bibliography = extract_bibliography(text)
    citations = extract_citations(bibliography)

    anio = extract_anio(text)
    categoria = extract_categoria(text)
    tipo = extract_tipo(text)
    acronimo = extract_acronimo(text)
    paginas = extract_paginas(text) 
    obs = extract_obs(text)
    resumen = extract_resumen(text)
    enlace = extract_enlace(text)
    cita = extract_cita(text)
    autores = ""

    doc = {
        "filename": file.filename,
        "title": title,
        "abstract": abstract,
        "keywords": keywords,
        "bibliography": bibliography,
        "citations": citations,
        "autores": autores,
        "Year": anio,
        "Category": categoria,
        "Type": tipo,
        "Acronym": acronimo,
        "Cites": "",
        "Pag": paginas,
        "Obs": obs,
        "Summary": resumen,
        "link": enlace,
        "citation": cita
    }
    os.remove(temp_path)
    return redirect(url_for('documents.home'))
    #no funciona
    #return render_template('complete_document.html', document=doc)

@uploads_bp.route('/upload_folder', methods=['GET', 'POST'])
def upload_folder():
    if request.method == 'POST':
        files = request.files.getlist('pdfs')
        if not files or len(files) == 0:
            flash("No se seleccionaron archivos.")
            return redirect(url_for('upload.upload_folder'))

        drafts = []
        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)

        for file in files:
            if file.filename == '' or not file.filename.lower().endswith('.pdf'):
                continue
            temp_path = os.path.join(temp_folder, file.filename)
            file.save(temp_path)

            text = extract_text_from_pdf(temp_path)
            title = extract_title(text)
            abstract = extract_abstract(text)
            keywords = extract_keywords(text)
            bibliography = extract_bibliography(text)
            citations = extract_citations(bibliography)

            anio = extract_anio(text)
            categoria = extract_categoria(text)
            tipo = extract_tipo(text)
            acronimo = extract_acronimo(text)
            paginas = extract_paginas(text)
            obs = extract_obs(text)
            resumen = extract_resumen(text)
            enlace = extract_enlace(text)
            cita = extract_cita(text)

            doc = {
                "filename": file.filename,
                "title": title,
                "abstract": abstract,
                "keywords": keywords,
                "bibliography": bibliography,
                "citations": citations,
                "Year": anio,
                "Category": categoria,
                "Type": tipo,
                "Acronym": acronimo,
                "Cites": "",
                "Pag": paginas,
                "Obs": obs,
                "Summary": resumen,
                "link": enlace,
                "citation": cita
            }
            drafts.append(doc)
            os.remove(temp_path)

        if drafts:
            batch_id = save_drafts_list(drafts)
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
        flash("No se han subido documentos válidos.")
        return redirect(url_for('upload.upload_page'))
    return render_template('upload_folder.html')





""""Upload Excel inicial"""

@uploads_bp.route('/upload_excel', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        files = request.files.getlist('excel') if 'excel' in request.files else []
        if not files:
            flash("No se encontró el archivo Excel.")
            return redirect(url_for('upload.upload_excel'))

        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)

        drafts = []
        required_columns = [
            "Year", "Title", "Category", "Type", "Acronym", "Cites",
            "Pag", "Obs", "Summary", "link", "citation", "abstract"
        ]

        from openpyxl import load_workbook
        for file in files:
            if not file or file.filename == '':
                continue
            temp_path = os.path.join(temp_folder, file.filename)
            file.save(temp_path)
            try:
                wb = load_workbook(temp_path)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                for col_name in required_columns:
                    if col_name not in headers:
                        flash(f"No se encontró la columna requerida: {col_name} en {file.filename}.")
                        # saltar este archivo
                        raise ValueError("Faltan columnas requeridas")
                col_index_map = { col_name: headers.index(col_name) + 1 for col_name in required_columns }
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    doc = {}
                    for col_name in required_columns:
                        idx = col_index_map[col_name] - 1
                        valor = row[idx] if idx < len(row) else None
                        doc[col_name] = valor if valor is not None else ""
                    drafts.append(doc)
            except Exception as e:
                # Reportar error y continuar con otros archivos
                flash(f"Error procesando {file.filename}: {str(e)}")
            finally:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass

        if drafts:
            batch_id = save_drafts_list(drafts)
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
        flash("No se han subido documentos desde los Excel seleccionados.")
        return redirect(url_for('upload.upload_page'))
    return render_template('upload_excel.html')

@uploads_bp.route('/upload_zip', methods=['POST'])
def upload_zip():
    if 'zipfile' not in request.files:
        flash("No se encontró el archivo ZIP.")
        return redirect(url_for('upload.upload_page'))

    zf = request.files['zipfile']
    if not zf or zf.filename == '':
        flash("No se seleccionó ningún ZIP.")
        return redirect(url_for('upload.upload_page'))

    temp_dir = tempfile.mkdtemp(prefix='upload_zip_')
    zip_path = os.path.join(temp_dir, zf.filename)
    zf.save(zip_path)

    pdf_docs = []
    excel_docs = []
    drafts = []
    try:
        with zipfile.ZipFile(zip_path, 'r') as z:
            members = [m for m in z.namelist() if not m.endswith('/')]
            for m in members:
                lower = m.lower()
                if lower.endswith('.pdf') or lower.endswith('.xlsx') or lower.endswith('.xls'):
                    out = os.path.join(temp_dir, os.path.basename(m))
                    with z.open(m) as src, open(out, 'wb') as dst:
                        dst.write(src.read())
                    if lower.endswith('.pdf'):
                        pdf_docs.append(out)
                    else:
                        excel_docs.append(out)

        if pdf_docs:
            for temp_path in pdf_docs:
                try:
                    text = extract_text_from_pdf(temp_path)
                    title = extract_title(text)
                    abstract = extract_abstract(text)
                    keywords = extract_keywords(text)
                    bibliography = extract_bibliography(text)
                    citations = extract_citations(bibliography)
                    doc = {
                        "filename": os.path.basename(temp_path),
                        "title": title,
                        "abstract": abstract,
                        "keywords": keywords,
                        "bibliography": bibliography,
                        "citations": citations,
                        "Year": extract_anio(text),
                        "Category": extract_categoria(text),
                        "Type": extract_tipo(text),
                        "Acronym": extract_acronimo(text),
                        "Cites": "",
                        "Pag": extract_paginas(text),
                        "Obs": extract_obs(text),
                        "Summary": extract_resumen(text),
                        "link": extract_enlace(text),
                        "citation": extract_cita(text)
                    }
                    drafts.append(doc)
                except Exception as e:
                    flash(f"Error procesando PDF {os.path.basename(temp_path)}: {str(e)}")

            if drafts:
                batch_id = save_drafts_list(drafts)
                return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
            flash("No se han subido documentos desde el ZIP.")
            return redirect(url_for('upload.upload_page'))

        if excel_docs:
            from openpyxl import load_workbook
            required_columns = [
                "Year", "Title", "Category", "Type", "Acronym", "Cites",
                "Pag", "Obs", "Summary", "link", "citation", "abstract"
            ]
            
            headers = [cell.value for cell in sheet[1]]
            for col_name in required_columns:
                if col_name not in headers:
                    flash(f"No se encontró la columna requerida: {col_name}.")
                    return redirect(url_for('documents.upload_excel'))
            
            col_index_map = { col_name: headers.index(col_name) + 1 for col_name in required_columns }
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                doc = {}
                for col_name in required_columns:
                    idx = col_index_map[col_name] - 1
                    valor = row[idx] if idx < len(row) else None
                    doc[col_name] = valor if valor is not None else ""
                mongo.db.documents.insert_one(doc)
                processed_docs.append(doc)
            
            flash(f"Se han subido {len(processed_docs)} documentos desde el Excel.")
        except Exception as e:
            flash(f"Ocurrió un error procesando el Excel: {str(e)}")
        finally:
            os.remove(temp_path)
        
        return render_template('result_excel.html', documents=processed_docs)
    
    return render_template('upload_excel.html')





############################################################
############################################################
    # RAMA FEATURE-BRANCH #
############################################################
############################################################


'''

# routes/uploads.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from bson.objectid import ObjectId
from extensions import mongo
from utils import (
    extract_text_from_pdf, extract_title, extract_abstract,
    extract_keywords, extract_bibliography, extract_citations,
    extract_anio, extract_categoria, extract_tipo, extract_acronimo,
    extract_paginas, extract_obs, extract_resumen, extract_enlace, extract_cita
)
import os
import zipfile
import tempfile
import uuid
import json
from routes.documents import edit_document as documents_edit_document

uploads_bp = Blueprint('upload', __name__, template_folder='../templates')

# --- Drafts helpers (store drafts as JSON files under temp/drafts) ---
DRAFTS_DIR = os.path.join("temp", "drafts")

def _ensure_drafts_dir():
    os.makedirs(DRAFTS_DIR, exist_ok=True)

def _drafts_path(batch_id: str) -> str:
    _ensure_drafts_dir()
    safe = f"{batch_id}.json"
    return os.path.join(DRAFTS_DIR, safe)

def save_drafts_list(drafts: list, batch_id: str | None = None) -> str:
    if batch_id is None:
        batch_id = uuid.uuid4().hex
    path = _drafts_path(batch_id)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(drafts, f, ensure_ascii=False)
    except Exception:
        # Fallback: ensure directory and retry once
        _ensure_drafts_dir()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(drafts, f, ensure_ascii=False)
    return batch_id

def load_drafts_list(batch_id: str) -> list:
    path = _drafts_path(batch_id)
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f) or []
    except Exception:
        return []

def update_draft(batch_id: str, index: int, draft: dict) -> None:
    drafts = load_drafts_list(batch_id)
    if 0 <= index < len(drafts):
        drafts[index] = draft
        save_drafts_list(drafts, batch_id=batch_id)

def cleanup_drafts(batch_id: str) -> None:
    path = _drafts_path(batch_id)
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass

@uploads_bp.route('/upload')
def upload_page():
    """Página para subir un PDF."""
    return render_template('upload.html')

# Mantener la URL de edición como /upload/<id> para 'completar documento'
@uploads_bp.route('/upload/<id>', methods=['GET', 'POST'])
def upload_edit(id):
    # Reutiliza la lógica de edición existente manteniendo la URL
    return documents_edit_document(id)

# Editor de borradores: inserta en MongoDB al pulsar Guardar/Siguiente
@uploads_bp.route('/upload/draft/<batch_id>/<int:index>', methods=['GET', 'POST'])
def draft_edit(batch_id, index: int):
    drafts = load_drafts_list(batch_id)
    if not drafts or index < 0 or index >= len(drafts):
        flash("Borrador no encontrado o expirado.")
        return redirect(url_for('upload.upload_page'))

    if request.method == 'POST':
        # Recoger campos estándar
        fields = [
            "Year", "Title", "Category", "Type", "Acronym", "Cites",
            "Pag", "Obs", "Summary", "link", "citation", "abstract", "autores", "filename",
            # También conservar variantes en minúscula si vienen del extractor
            "title", "keywords", "bibliography", "citations"
        ]
        updated = dict(drafts[index])
        for f in fields:
            val = request.form.get(f, None)
            if val is not None:
                updated[f] = val

        # Persistir cambios en el borrador
        update_draft(batch_id, index, updated)

        # Normalizar título antes de insertar
        try:
            t_cap = (updated.get('Title') or '').strip()
            t_low = (updated.get('title') or '').strip()
            if t_cap:
                updated['Title'] = t_cap
                updated['title'] = t_cap
            elif t_low:
                updated['Title'] = t_low
                updated['title'] = t_low
        except Exception:
            pass

        # Insertar en Mongo al guardar
        try:
            ins = mongo.db.documents.insert_one(updated)
            inserted_id = str(ins.inserted_id)
        except Exception as e:
            flash(f"Error guardando: {str(e)}")
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=index))

        # Navegación por lote (ir al siguiente borrador si procede)
        in_batch = request.form.get('batch', request.args.get('batch', '0')) == '1'
        go_next = request.form.get('next', '0') == '1'

        if in_batch and go_next and (index + 1) < len(drafts):
            # Continuar editando el siguiente borrador sin ID todavía
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=index+1, batch=1))

        # Si es el último o no hay flujo de siguiente, limpiar borradores
        if (index >= len(drafts) - 1) or not in_batch:
            cleanup_drafts(batch_id)

        # Tras crear el documento real, opcionalmente respetar URL de retorno
        ret = request.form.get('return', '').strip()
        if ret:
            try:
                from urllib.parse import urlparse
                host = urlparse(request.host_url)
                target = urlparse(ret)
                if target.scheme in ('http','https') and target.netloc == host.netloc:
                    return redirect(ret)
            except Exception:
                pass
        # Si no hay retorno, continuar en /upload/<id> para futuras ediciones
        return redirect(url_for('upload.upload_edit', id=inserted_id))

    # GET: Mostrar el borrador en el editor
    doc = drafts[index]
    return render_template(
        'complete_document.html',
        document=doc,
        batch=(len(drafts) > 1),
        index=index,
        is_last=(index >= len(drafts) - 1),
        draft_batch=batch_id
    )

@uploads_bp.route('/upload_pdf', methods=['POST'])
def upload_pdf():
    if 'pdf' not in request.files:
        flash("No se encontró el archivo PDF.")
        return redirect(url_for('upload.upload_page'))
    file = request.files['pdf']
    if file.filename == '':
        flash("No se seleccionó ningún archivo.")
        return redirect(url_for('upload.upload_page'))
    autores = request.form.get('autores', '').strip()
    anio = request.form.get('anio', '').strip()
    categoria = request.form.get('categoria', '').strip()
    tipo = request.form.get('tipo', '').strip()
    acronimo = request.form.get('acronimo', '').strip()
    paginas = request.form.get('paginas', '').strip()
    obs = request.form.get('obs', '').strip()  
    resumen = request.form.get('resumen', '').strip()
    enlace = request.form.get('enlace', '').strip()
    cita = request.form.get('cita', '').strip()
    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)
    temp_path = os.path.join(temp_folder, file.filename)
    file.save(temp_path)
    text = extract_text_from_pdf(temp_path)
    title = extract_title(text)
    abstract = extract_abstract(text)
    keywords = extract_keywords(text)
    bibliography = extract_bibliography(text)
    citations = extract_citations(bibliography)
    anio = extract_anio(text) if not anio else anio
    categoria = extract_categoria(text) if not categoria else categoria
    tipo = extract_tipo(text) if not tipo else tipo
    acronimo = extract_acronimo(text) if not acronimo else acronimo
    paginas = extract_paginas(text) if not paginas else paginas
    obs = extract_obs(text) if not obs else obs
    resumen = extract_resumen(text) if not resumen else resumen
    enlace = extract_enlace(text) if not enlace else enlace
    cita = extract_cita(text) if not cita else cita
    doc = {
        "filename": file.filename,
        "title": title,
        "abstract": abstract,
        "keywords": keywords,
        "bibliography": bibliography,
        "citations": citations,
        "autores": autores,
        "Year": anio,
        "Category": categoria,
        "Type": tipo,
        "Acronym": acronimo,
        "Cites": "",
        "Pag": paginas,
        "Obs": obs,
        "Summary": resumen,
        "link": enlace,
        "citation": cita
    }
    os.remove(temp_path)
    batch_id = save_drafts_list([doc])
    return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=0))






"""Upload Folder feature-branch"""

@uploads_bp.route('/upload_folder', methods=['GET', 'POST'])
def upload_folder():
    if request.method == 'POST':
        files = request.files.getlist('pdfs')
        if not files or len(files) == 0:
            flash("No se seleccionaron archivos.")
            return redirect(url_for('upload.upload_folder'))

        drafts = []
        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)

        for file in files:
            if file.filename == '' or not file.filename.lower().endswith('.pdf'):
                continue
            temp_path = os.path.join(temp_folder, file.filename)
            file.save(temp_path)

            text = extract_text_from_pdf(temp_path)
            title = extract_title(text)
            abstract = extract_abstract(text)
            keywords = extract_keywords(text)
            bibliography = extract_bibliography(text)
            citations = extract_citations(bibliography)

            anio = extract_anio(text)
            categoria = extract_categoria(text)
            tipo = extract_tipo(text)
            acronimo = extract_acronimo(text)
            paginas = extract_paginas(text)
            obs = extract_obs(text)
            resumen = extract_resumen(text)
            enlace = extract_enlace(text)
            cita = extract_cita(text)

            doc = {
                "filename": file.filename,
                "title": title,
                "abstract": abstract,
                "keywords": keywords,
                "bibliography": bibliography,
                "citations": citations,
                "Year": anio,
                "Category": categoria,
                "Type": tipo,
                "Acronym": acronimo,
                "Cites": "",
                "Pag": paginas,
                "Obs": obs,
                "Summary": resumen,
                "link": enlace,
                "citation": cita
            }
            drafts.append(doc)
            os.remove(temp_path)

        if drafts:
            batch_id = save_drafts_list(drafts)
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
        flash("No se han subido documentos válidos.")
        return redirect(url_for('upload.upload_page'))
    return render_template('upload_folder.html')



"""Upload Excel feature-branch"""

@uploads_bp.route('/upload_excel', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        files = request.files.getlist('excel') if 'excel' in request.files else []
        if not files:
            flash("No se encontró el archivo Excel.")
            return redirect(url_for('upload.upload_excel'))

        temp_folder = os.path.join("temp")
        os.makedirs(temp_folder, exist_ok=True)

        drafts = []
        required_columns = [
            "Year", "Title", "Category", "Type", "Acronym", "Cites",
            "Pag", "Obs", "Summary", "link", "citation", "abstract"
        ]

        from openpyxl import load_workbook
        for file in files:
            if not file or file.filename == '':
                continue
            temp_path = os.path.join(temp_folder, file.filename)
            file.save(temp_path)
            try:
                wb = load_workbook(temp_path)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                for col_name in required_columns:
                    if col_name not in headers:
                        flash(f"No se encontró la columna requerida: {col_name} en {file.filename}.")
                        # saltar este archivo
                        raise ValueError("Faltan columnas requeridas")
                col_index_map = { col_name: headers.index(col_name) + 1 for col_name in required_columns }
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    doc = {}
                    for col_name in required_columns:
                        idx = col_index_map[col_name] - 1
                        valor = row[idx] if idx < len(row) else None
                        doc[col_name] = valor if valor is not None else ""
                    drafts.append(doc)
            except Exception as e:
                # Reportar error y continuar con otros archivos
                flash(f"Error procesando {file.filename}: {str(e)}")
            finally:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass

        if drafts:
            batch_id = save_drafts_list(drafts)
            return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
        flash("No se han subido documentos desde los Excel seleccionados.")
        return redirect(url_for('upload.upload_page'))
    return render_template('upload_excel.html')


"""Upload ZIP feature-branch"""

@uploads_bp.route('/upload_zip', methods=['POST'])
def upload_zip():
    if 'zipfile' not in request.files:
        flash("No se encontró el archivo ZIP.")
        return redirect(url_for('upload.upload_page'))

    zf = request.files['zipfile']
    if not zf or zf.filename == '':
        flash("No se seleccionó ningún ZIP.")
        return redirect(url_for('upload.upload_page'))

    temp_dir = tempfile.mkdtemp(prefix='upload_zip_')
    zip_path = os.path.join(temp_dir, zf.filename)
    zf.save(zip_path)

    pdf_docs = []
    excel_docs = []
    drafts = []
    try:
        with zipfile.ZipFile(zip_path, 'r') as z:
            members = [m for m in z.namelist() if not m.endswith('/')]
            for m in members:
                lower = m.lower()
                if lower.endswith('.pdf') or lower.endswith('.xlsx') or lower.endswith('.xls'):
                    out = os.path.join(temp_dir, os.path.basename(m))
                    with z.open(m) as src, open(out, 'wb') as dst:
                        dst.write(src.read())
                    if lower.endswith('.pdf'):
                        pdf_docs.append(out)
                    else:
                        excel_docs.append(out)

        if pdf_docs:
            for temp_path in pdf_docs:
                try:
                    text = extract_text_from_pdf(temp_path)
                    title = extract_title(text)
                    abstract = extract_abstract(text)
                    keywords = extract_keywords(text)
                    bibliography = extract_bibliography(text)
                    citations = extract_citations(bibliography)
                    doc = {
                        "filename": os.path.basename(temp_path),
                        "title": title,
                        "abstract": abstract,
                        "keywords": keywords,
                        "bibliography": bibliography,
                        "citations": citations,
                        "Year": extract_anio(text),
                        "Category": extract_categoria(text),
                        "Type": extract_tipo(text),
                        "Acronym": extract_acronimo(text),
                        "Cites": "",
                        "Pag": extract_paginas(text),
                        "Obs": extract_obs(text),
                        "Summary": extract_resumen(text),
                        "link": extract_enlace(text),
                        "citation": extract_cita(text)
                    }
                    drafts.append(doc)
                except Exception as e:
                    flash(f"Error procesando PDF {os.path.basename(temp_path)}: {str(e)}")

            if drafts:
                batch_id = save_drafts_list(drafts)
                return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
            flash("No se han subido documentos desde el ZIP.")
            return redirect(url_for('upload.upload_page'))

        if excel_docs:
            from openpyxl import load_workbook
            required_columns = [
                "Year", "Title", "Category", "Type", "Acronym", "Cites",
                "Pag", "Obs", "Summary", "link", "citation", "abstract"
            ]
            for temp_path in excel_docs:
                try:
                    wb = load_workbook(temp_path)
                    sheet = wb.active
                    headers = [cell.value for cell in sheet[1]]
                    for col_name in required_columns:
                        if col_name not in headers:
                            flash(f"No se encontró la columna requerida: {col_name} en {os.path.basename(temp_path)}.")
                            raise ValueError("Faltan columnas requeridas")
                    col_index_map = { col_name: headers.index(col_name) + 1 for col_name in required_columns }
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not any(row):
                            continue
                        doc = {}
                        for col_name in required_columns:
                            idx = col_index_map[col_name] - 1
                            valor = row[idx] if idx < len(row) else None
                            doc[col_name] = valor if valor is not None else ""
                        drafts.append(doc)
                except Exception as e:
                    flash(f"Error procesando Excel {os.path.basename(temp_path)}: {str(e)}")

            if drafts:
                batch_id = save_drafts_list(drafts)
                return redirect(url_for('upload.draft_edit', batch_id=batch_id, index=0, batch=1))
            flash("No se han subido documentos desde el ZIP.")
            return redirect(url_for('upload.upload_page'))

        flash("El ZIP no contiene PDFs ni Excels válidos.")
        return redirect(url_for('upload.upload_page'))
    finally:
        try:
            for root, dirs, files in os.walk(temp_dir, topdown=False):
                for name in files:
                    try:
                        os.remove(os.path.join(root, name))
                    except OSError:
                        pass
                for name in dirs:
                    try:
                        os.rmdir(os.path.join(root, name))
                    except OSError:
                        pass
            os.rmdir(temp_dir)
        except OSError:
            pass

'''